from flask import Flask, render_template, jsonify, send_file, request
import pandas as pd
import os
from datetime import datetime
import threading
import requests
import xml.etree.ElementTree as ET
from dotenv import load_dotenv
from scraper import scrape_ftc_law_data
from urllib.parse import unquote, quote

# .env 파일 로드
load_dotenv()

app = Flask(__name__)

# 전역 변수로 진행 상황 및 데이터 저장
scraping_status = {
    "is_running": False,
    "progress": 0,
    "total": 0,
    "current_category": "",
    "data": [],
    "pdf_zip_path": None,
    "target_dir": None
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/scrape/start', methods=['POST'])
def start_scrape():
    global scraping_status
    if scraping_status["is_running"]:
        return jsonify({"status": "error", "message": "이미 스크래핑이 진행 중입니다."})

    target_cd = request.json.get('target_cd', '1')
    scraping_status["is_running"] = True
    scraping_status["progress"] = 0
    scraping_status["data"] = []
    
    thread = threading.Thread(target=run_scraping_task, args=(target_cd,))
    thread.start()
    
    return jsonify({"status": "success", "message": "스크래핑을 시작합니다."})

def run_scraping_task(target_cd):
    global scraping_status
    all_dfs = []
    
    try:
        if target_cd == 'all':
            cds = [f"{i:02d}" for i in range(1, 15)]
        else:
            cds = [target_cd]
        
        scraping_status["total"] = 0
        scraping_status["progress"] = 0
        
        for idx, cd in enumerate(cds):
            cd_int = int(cd)
            scraping_status["current_category"] = f"카테고리 {cd} 수집 중... ({idx+1}/{len(cds)})"
            df = scrape_ftc_law_data(cd_int)
            if not df.empty:
                all_dfs.append(df)
                total_laws = sum(len(d) for d in all_dfs)
                scraping_status["total"] = total_laws
                scraping_status["progress"] = total_laws
            
        if all_dfs:
            combined_df = pd.concat(all_dfs, ignore_index=True)
            scraping_status["data"] = combined_df.to_dict('records')
            scraping_status["total"] = len(scraping_status["data"])
            scraping_status["progress"] = len(scraping_status["data"])
        
    except Exception as e:
        print(f"Scraping error: {e}")
    finally:
        scraping_status["is_running"] = False
        scraping_status["current_category"] = "완료"

@app.route('/api/scrape/info', methods=['POST'])
def scrape_info():
    global scraping_status
    if not scraping_status["data"]:
        return jsonify({"status": "error", "message": "먼저 데이터 수집을 완료해주세요."})
    
    if scraping_status["is_running"]:
        return jsonify({"status": "error", "message": "이미 다른 작업이 진행 중입니다."})

    scraping_status["is_running"] = True
    scraping_status["progress"] = 0
    scraping_status["total"] = len(scraping_status["data"])
    
    thread = threading.Thread(target=run_info_update_task)
    thread.start()
    
    return jsonify({"status": "success", "message": "API를 이용한 시행/개정 정보 수집을 시작합니다."})

def run_info_update_task():
    global scraping_status
    import time
    from urllib.parse import unquote, quote
    import re
    
    # 1. API 키 로드 (공백 제거)
    api_personal = os.environ.get('LAW_API_KEY_PERSONAL', '').strip()
    api_corporate = os.environ.get('LAW_API_KEY_CORPORATE', '').strip()
    api_key_raw = os.environ.get('DATA_GO_KR_API_KEY', '').strip() or api_personal or api_corporate
    
    if not api_key_raw or "your_" in api_key_raw:
        scraping_status["is_running"] = False
        scraping_status["current_category"] = "오류: API 키를 설정해주세요 (.env 파일)"
        return

    # 인증키 처리 (이미 인코딩된 경우 대응)
    api_key = unquote(api_key_raw)
    
    # ServiceKey 여부 판별 고도화 (Base64 특성 및 길이 확인)
    # data.go.kr 서비스키는 보통 +, /, = 등이 포함되거나 60자 이상의 긴 문자열임
    is_base64 = bool(re.search(r'[+/=]', api_key)) or len(api_key) > 80
    
    data_list = scraping_status["data"]
    
    print(f"INFO: API Update Task Started. Target: {len(data_list)} items. Detected Type: {'ServiceKey' if is_base64 else 'OC ID'}")

    for i in range(len(data_list)):
        item = data_list[i]
        name = item.get('법령명_상세', '')
        category_type = item.get('구분', '')
        
        if not name:
            scraping_status["progress"] = i + 1
            continue

        scraping_status["current_category"] = f"[{i+1}/{len(data_list)}] {name} API 조회 중..."
        
        target = "eflaw"
        if any(x in category_type for x in ["고시", "지침", "예규", "훈령", "공고"]):
            target = "admrul"
            
        success = False
        # 시도할 설정들 (공공데이터포털 vs 국가법령정보센터)
        configs = []
        if is_base64:
            configs = [
                {"url": "http://apis.data.go.kr/1170000/law/lawSearchList.do", "param": "serviceKey"},
                {"url": "http://www.law.go.kr/DRF/lawSearch.do", "param": "OC"}
            ]
        else:
            configs = [
                {"url": "http://www.law.go.kr/DRF/lawSearch.do", "param": "OC"},
                {"url": "http://apis.data.go.kr/1170000/law/lawSearchList.do", "param": "serviceKey"}
            ]

        for cfg in configs:
            try:
                params = {
                    cfg["param"]: api_key,
                    "target": target,
                    "type": "XML",
                    "query": name,
                    "mobileYn": "Y"
                }
                
                response = requests.get(cfg["url"], params=params, timeout=10, verify=False)
                
                if response.status_code == 200:
                    content = response.content.decode('utf-8', errors='replace')
                    
                    # 인증 오류 키워드 체크
                    if any(msg in content for msg in ["인증되지 않은", "ErrorCode", "SERVICE_KEY_IS_NOT_REGISTERED"]):
                        print(f"DEBUG: Auth failure on {cfg['url']} for {name}, trying next config...")
                        continue

                    # 정상 XML 파싱 시도
                    try:
                        root = ET.fromstring(content.encode('utf-8'))
                        law_node = root.find(f".//{target}")
                        
                        if law_node is not None:
                            impl_date = law_node.findtext("시행일자") or law_node.findtext("발령일자") or ""
                            if len(impl_date) == 8:
                                impl_date = f"{impl_date[:4]}. {impl_date[4:6]}. {impl_date[6:8]}."
                            
                            pnt_no = law_node.findtext("공포번호") or law_node.findtext("발령번호") or ""
                            pnt_date = law_node.findtext("공포일자") or law_node.findtext("발령일자") or ""
                            if len(pnt_date) == 8:
                                pnt_date = f"{pnt_date[:4]}. {pnt_date[4:6]}. {pnt_date[6:8]}."
                            
                            cat_name = law_node.findtext("법령구분명") or law_node.findtext("행정규칙종류명") or category_type
                            rev_name = law_node.findtext("제개정구분명") or "-"
                            
                            full_info = f"[시행 {impl_date}] [{cat_name} 제{pnt_no}호, {pnt_date}, {rev_name}]"
                            
                            data_list[i].update({
                                "시행/개정": full_info,
                                "시행일": impl_date or "-",
                                "개정유형": rev_name,
                                "개정정보": f"{cat_name} 제{pnt_no}호" if pnt_no else "-",
                                "개정일": pnt_date or "-"
                            })
                            success = True
                            break # 성공 시 다음 설정 시도 중단
                    except:
                        continue # 파싱 에러 시 다음 설정 시도
                else:
                    # 401/403 등 오류 시 다음 설정 시도
                    print(f"DEBUG: HTTP {response.status_code} on {cfg['url']}, trying next...")
                    continue

            except Exception as e:
                print(f"DEBUG: Connection error on {cfg['url']}: {e}")
                continue

        if not success:
            # 모든 시도가 실패한 경우
            data_list[i].update({
                "시행/개정": "인증/조회 실패",
                "시행일": "API 오류 (키 확인)",
                "개정유형": "-", "개정정보": "-", "개정일": "-"
            })
            
        scraping_status["progress"] = i + 1
        time.sleep(0.1)

    scraping_status["is_running"] = False
    scraping_status["current_category"] = "API 수집 완료"

@app.route('/api/scrape/status')
def get_status():
    resp = jsonify({
        "is_running": scraping_status["is_running"],
        "progress": scraping_status["progress"],
        "total": scraping_status["total"],
        "current_category": scraping_status["current_category"],
        "data_count": len(scraping_status["data"])
    })
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

@app.route('/api/scrape/results')
def get_results():
    resp = jsonify(scraping_status["data"])
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    if not scraping_status["data"]:
        return jsonify({"status": "error", "message": "저장할 데이터가 없습니다."})
        
    df = pd.DataFrame(scraping_status["data"])
    cols = ["법령명", "구분", "법령명_상세", "담당부서", "개정유형", "시행일", "개정정보", "개정일", "팝업페이지링크"]
    df = df.reindex(columns=cols, fill_value="-")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"FTC_Laws_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='FTC_Laws')
            workbook = writer.book
            worksheet = writer.sheets['FTC_Laws']
            
            header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
            
            col_widths = {'A': 15, 'B': 12, 'C': 35, 'D': 15, 'E': 12, 'F': 15, 'G': 20, 'H': 15, 'I': 20}
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter in ['C', 'G', 'I']:
                        cell.alignment = alignment_left
                    else:
                        cell.alignment = alignment_center
    except:
        df.to_excel(filepath, index=False)
    
    return jsonify({"status": "success", "filename": filename})

@app.route('/api/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(OUTPUT_DIR, filename), as_attachment=True)

@app.route('/api/pdf/save', methods=['POST'])
def save_pdf():
    if not scraping_status["data"]:
        return jsonify({"status": "error", "message": "저장할 데이터가 없습니다."})
    
    if scraping_status["is_running"]:
        return jsonify({"status": "error", "message": "현재 다른 작업이 진행 중입니다."})

    scraping_status["is_running"] = True
    scraping_status["progress"] = 0
    scraping_status["total"] = len(scraping_status["data"])
    
    thread = threading.Thread(target=run_pdf_save_task)
    thread.start()
    
    return jsonify({"status": "success", "message": "PDF 저장을 시작합니다."})

@app.route('/api/pdf/download')
def download_pdf_zip():
    zip_path = scraping_status.get("pdf_zip_path")
    if not zip_path or not os.path.exists(zip_path):
        return jsonify({"status": "error", "message": "다운로드할 PDF 파일이 없습니다."})
    return send_file(zip_path, as_attachment=True)

@app.route('/api/pdf/status')
def get_pdf_status():
    return jsonify({
        "has_zip": bool(scraping_status.get("pdf_zip_path") and os.path.exists(scraping_status.get("pdf_zip_path", "")))
    })

def run_pdf_save_task():
    global scraping_status
    import shutil
    from playwright.sync_api import sync_playwright
    
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_dir = os.path.join(OUTPUT_DIR, ts)
    os.makedirs(pdf_dir, exist_ok=True)
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            
            for i, item in enumerate(scraping_status["data"]):
                url = item.get("팝업페이지링크")
                if url and url.startswith("http"):
                    name = item.get("법령명_상세", f"law_{i}")
                    safe_name = "".join([c for c in name if c.isalnum() or c in (' ', '_')]).strip()
                    pdf_path = os.path.join(pdf_dir, f"{safe_name}.pdf")
                    
                    try:
                        page = context.new_page()
                        page.goto(url, wait_until="networkidle", timeout=60000)
                        page.pdf(path=pdf_path)
                        page.close()
                    except:
                        pass
                scraping_status["progress"] = i + 1
            
            browser.close()
            
            zip_path = os.path.join(OUTPUT_DIR, f"FTC_Laws_PDF_{ts}")
            shutil.make_archive(zip_path, 'zip', pdf_dir)
            scraping_status["pdf_zip_path"] = zip_path + ".zip"
            
    except Exception as e:
        print(f"PDF save error: {e}")
    finally:
        scraping_status["is_running"] = False
        scraping_status["current_category"] = "PDF 저장 완료"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
