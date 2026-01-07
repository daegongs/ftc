import openpyxl

def extract_links_to_column_d(filename):
    print(f"[INFO] '{filename}' 파일을 불러오는 중입니다...")
    
    try:
        # 1. 엑셀 파일 불러오기
        wb = openpyxl.load_workbook(filename)
        ws = wb.active  # 활성화된 시트 선택
        
        count = 0 # 추출된 링크 개수를 세기 위한 변수

        # 2. 데이터가 있는 행 반복 (2번째 줄부터 시작)
        # min_row=2: 헤더 제외
        # min_col=2: B열(법령명)
        # max_col=2: B열만 확인
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            cell = row[0] # iter_rows는 튜플을 반환하므로 첫 번째 요소가 해당 셀입니다.
            
            # 3. 하이퍼링크가 존재하는지 확인
            if cell.hyperlink:
                target_url = cell.hyperlink.target
                
                # 4. D열(4번째 열)에 링크 주소 입력
                # cell.row는 현재 행 번호, column=4는 D열
                ws.cell(row=cell.row, column=4).value = target_url
                count += 1

        # 5. 결과 저장하기
        new_filename = filename.replace('.xlsx', '_result.xlsx')
        wb.save(new_filename)
        
        print(f"[OK] 작업 완료! 총 {count}개의 링크를 추출했습니다.")
        print(f"[OK] 결과 파일이 '{new_filename}'으로 저장되었습니다.")

    except FileNotFoundError:
        print("[ERROR] 파일을 찾을 수 없습니다. 파일명이 정확한지 확인해 주세요.")
    except Exception as e:
        print(f"[ERROR] 오류 발생: {e}")

# --- 실행부 ---
# 원본 엑셀 파일명을 여기에 적어주세요
target_file = 'ftc_law_data.xlsx' 

extract_links_to_column_d(target_file)