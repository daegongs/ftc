import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import re
import os

def get_driver():
    """Chrome WebDriver 설정"""
    chrome_options = Options()
    chrome_options.add_argument('--headless=new')  # 새로운 headless 모드
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    try:
        # webdriver-manager 사용 시도
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except:
            # webdriver-manager 실패 시 직접 시도
            driver = webdriver.Chrome(options=chrome_options)
        
        driver.set_page_load_timeout(30)
        print("[OK] ChromeDriver 초기화 성공", flush=True)
        return driver
    except Exception as e:
        error_msg = str(e)
        print(f"[ERROR] ChromeDriver 설정 실패: {error_msg}", flush=True)
        
        if "chromedriver" in error_msg.lower() or "webdriver" in error_msg.lower():
            print("\n[해결 방법]", flush=True)
            print("1. Chrome 브라우저가 설치되어 있는지 확인", flush=True)
            print("2. 다음 명령어로 webdriver-manager 재설치:", flush=True)
            print("   py -m pip install --upgrade webdriver-manager", flush=True)
            print("3. 또는 ChromeDriver를 수동으로 다운로드:", flush=True)
            print("   https://googlechromelabs.github.io/chrome-for-testing/", flush=True)
        
        return None

def extract_implementation_date(driver, url):
    """URL에서 시행일시 추출"""
    try:
        # HTTP를 HTTPS로 변경
        if url.startswith('http://'):
            url = url.replace('http://', 'https://')
        
        # 페이지 접속
        driver.get(url)
        
        # 페이지 로드 대기
        time.sleep(2)
        
        # iframe 찾기 및 전환
        try:
            # iframe이 로드될 때까지 대기
            wait = WebDriverWait(driver, 10)
            iframe = wait.until(EC.presence_of_element_located((By.ID, "lawService")))
            
            # iframe으로 전환
            driver.switch_to.frame(iframe)
            
            # iframe 내부 콘텐츠 로드 대기
            time.sleep(2)
            
            # 페이지 소스 가져오기
            iframe_html = driver.page_source
            iframe_soup = BeautifulSoup(iframe_html, 'html.parser')
            
            extracted_text = ""
            
            # 방법 1: span.tx2 (법률용)
            target_element = iframe_soup.select_one('span.tx2')
            if target_element:
                extracted_text = target_element.get_text(strip=True)
            
            # 방법 2: 행정규칙용 - '시행' 텍스트가 포함된 요소 찾기
            if not extracted_text:
                for tag in iframe_soup.find_all(string=lambda t: t and '시행' in t and len(t.strip()) < 200):
                    text = tag.strip()
                    # 날짜 형식이 포함된 경우
                    if any(char.isdigit() for char in text):
                        extracted_text = text
                        break
            
            # 방법 3: 정규식으로 날짜 패턴 검색
            if not extracted_text:
                all_text = iframe_soup.get_text()
                date_patterns = [
                    r'\d{4}\s*\.\s*\d{1,2}\s*\.\s*\d{1,2}',  # 2025. 1. 21.
                    r'\d{4}-\d{1,2}-\d{1,2}',  # 2025-01-21
                    r'\d{4}\.\s*\d{1,2}\.\s*\d{1,2}',  # 2025.1.21
                ]
                
                for pattern in date_patterns:
                    matches = re.findall(pattern, all_text)
                    if matches:
                        for match in matches:
                            idx = all_text.find(match)
                            if idx >= 0:
                                context = all_text[max(0, idx-50):idx+len(match)+50]
                                if '시행' in context or '고시' in context or '공고' in context:
                                    lines = context.split('\n')
                                    for line in lines:
                                        if match in line and ('시행' in line or '고시' in line or '공고' in line):
                                            extracted_text = line.strip()
                                            break
                                    if extracted_text:
                                        break
                        if extracted_text:
                            break
            
            # iframe에서 나오기
            driver.switch_to.default_content()
            
            return extracted_text
            
        except Exception as e:
            # iframe이 없는 경우 직접 페이지에서 검색
            driver.switch_to.default_content()
            page_html = driver.page_source
            page_soup = BeautifulSoup(page_html, 'html.parser')
            
            # span.tx2 찾기
            target_element = page_soup.select_one('span.tx2')
            if target_element:
                return target_element.get_text(strip=True)
            
            # '시행' 텍스트 검색
            for tag in page_soup.find_all(string=lambda t: t and '시행' in t and len(t.strip()) < 200):
                text = tag.strip()
                if any(char.isdigit() for char in text):
                    return text
            
            return ""
            
    except Exception as e:
        print(f"      [디버그] 추출 중 에러: {str(e)}", flush=True)
        return ""

def scrape_law_dates(filename):
    print(f"[INFO] '{filename}' 파일을 불러오는 중입니다...", flush=True)
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        print("[ERROR] 파일을 찾을 수 없습니다. 파일명이 정확한지 확인해주세요.", flush=True)
        return

    # E열 헤더 추가
    ws['E1'] = "시행일시"
    
    print("[INFO] 전체 행 처리를 시작합니다. (Selenium 사용 - 시간이 소요될 수 있습니다)", flush=True)
    
    # WebDriver 초기화
    driver = get_driver()
    if not driver:
        print("[ERROR] WebDriver를 초기화할 수 없습니다.", flush=True)
        return

    processed_count = 0
    success_count = 0
    fail_count = 0
    
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=5), start=2):
            processed_count += 1
            cell_b = row[1]  # B열 (법령명)
            cell_c = row[2]  # C열 (링크 - 하이퍼링크)
            
            law_name = cell_b.value
            target_url = ""

            # C열의 하이퍼링크 확인
            if cell_c.hyperlink:
                target_url = cell_c.hyperlink.target
            
            if target_url:
                try:
                    extracted_text = extract_implementation_date(driver, target_url)
                    
                    if extracted_text:
                        print(f"\n[{processed_count}/{row_idx}] {law_name}", flush=True)
                        print(f"  시행정보: {extracted_text}", flush=True)
                        ws.cell(row=row_idx, column=5).value = extracted_text
                        success_count += 1
                    else:
                        print(f"\n[{processed_count}/{row_idx}] {law_name}", flush=True)
                        print(f"  시행정보: (찾을 수 없음)", flush=True)
                        ws.cell(row=row_idx, column=5).value = "찾을 수 없음"
                        fail_count += 1
                    
                except Exception as e:
                    print(f"\n[{processed_count}/{row_idx}] {law_name}", flush=True)
                    print(f"  시행정보: (에러 - {str(e)[:50]})", flush=True)
                    ws.cell(row=row_idx, column=5).value = f"에러: {str(e)[:50]}"
                    fail_count += 1
            else:
                print(f"\n[{processed_count}/{row_idx}] {law_name}", flush=True)
                print(f"  시행정보: (링크 없음)", flush=True)
                ws.cell(row=row_idx, column=5).value = "링크 없음"
                fail_count += 1

            # 서버 부하 방지를 위해 1초 대기
            time.sleep(1)
    
    finally:
        # WebDriver 종료
        driver.quit()
        print("\n[INFO] WebDriver 종료", flush=True)

    # 결과 저장
    new_filename = "ftc_law_list_result.xlsx"
    wb.save(new_filename)
    
    # 최종 결과 출력
    print("\n" + "=" * 60, flush=True)
    print("[완료] 작업 완료!", flush=True)
    print("=" * 60, flush=True)
    print(f"처리된 항목: {processed_count}개", flush=True)
    print(f"성공: {success_count}개", flush=True)
    print(f"실패: {fail_count}개", flush=True)
    print(f"성공률: {(success_count/processed_count*100):.1f}%" if processed_count > 0 else "0%", flush=True)
    print(f"\n결과 파일: '{new_filename}'", flush=True)
    print("=" * 60, flush=True)

# 실행
if __name__ == "__main__":
    # 원본 엑셀 파일명: ftc_law_list.xlsx
    # 구조: A열(법령), B열(법령명), C열(링크), E열(시행정보 저장)
    scrape_law_dates("ftc_law_list.xlsx")
