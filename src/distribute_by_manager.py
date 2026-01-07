# -*- coding: utf-8 -*-
"""
담당자별 임원지분현황 배포 파일 생성 스크립트 (수정버전)
- 담당자정보 파일에서 담당자별 법인 매핑 정보 추출
- 임원지분현황 파일을 담당자별로 시트 분리
- 각 시트는 원본 양식(헤더 및 데이터 서식)과 동일하게 구성
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
import warnings

# 경고 메시지 숨김 (깔끔한 출력을 위해)
warnings.filterwarnings("ignore")


def normalize_company_name(name):
    """회사명 정규화: 공백 제거, 앞뒤 공백 제거"""
    if pd.isna(name):
        return ""
    s = str(name).strip()
    # 여러 공백을 하나로 및 모든 공백 제거 (비교 정확도 향상)
    return s.replace(" ", "")


def normalize_biz_number(num):
    """사업자번호 정규화: 하이픈 제거, 소수점 제거"""
    if pd.isna(num):
        return ""
    s = str(num).replace("-", "").strip()
    if "." in s:
        s = s.split(".")[0]
    return s


def load_manager_info(file_path):
    """담당자정보 파일 로드"""
    print(f"[1] 담당자정보 파일 로드 중: {file_path}")

    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = xl.sheet_names[0]
        df = pd.read_excel(
            file_path, sheet_name=sheet_name, dtype=str
        )  # 모든 데이터를 문자로 로드

        manager_col = None
        company_col = None
        biz_num_col = None

        for col in df.columns:
            col_lower = str(col).lower()
            if any(k in col_lower for k in ["담당자", "manager", "담당"]):
                manager_col = col
            if any(k in col_lower for k in ["법인", "회사", "company"]):
                company_col = col
            if any(k in col_lower for k in ["사업자", "biz", "등록번호"]):
                biz_num_col = col

        if not manager_col or not company_col:
            print("[ERROR] 담당자 또는 법인명 컬럼을 찾을 수 없습니다.")
            return None, None

        manager_companies = {}
        for _, row in df.iterrows():
            manager = (
                str(row[manager_col]).strip() if pd.notna(row[manager_col]) else ""
            )
            company = normalize_company_name(row[company_col])

            if manager and company:
                if manager not in manager_companies:
                    manager_companies[manager] = []

                company_info = {
                    "company_name": company,
                    "biz_number": (
                        normalize_biz_number(row[biz_num_col])
                        if biz_num_col and pd.notna(row[biz_num_col])
                        else ""
                    ),
                }
                if company_info not in manager_companies[manager]:
                    manager_companies[manager].append(company_info)

        print(f"    총 {len(manager_companies)}명의 담당자 정보 로드 완료")
        return manager_companies, df

    except Exception as e:
        print(f"[ERROR] 담당자정보 파일 로드 실패: {e}")
        raise


def load_stock_holding_data(file_path):
    """임원지분현황 파일 로드"""
    print(f"\n[2] 임원지분현황 파일 로드 중: {file_path}")

    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = xl.sheet_names[0]

        # 헤더 위치 찾기
        df_temp = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
        header_row = 0
        for idx, row in df_temp.iterrows():
            row_str = " ".join([str(v) for v in row if pd.notna(v)])
            if any(k in row_str for k in ["법인", "회사", "Company"]):
                header_row = idx
                break

        # 데이터 로드 (dtype=object로 설정하여 숫자 변형 방지)
        df = pd.read_excel(
            file_path, sheet_name=sheet_name, header=header_row, dtype=object
        )

        company_col = None
        biz_num_col = None

        for col in df.columns:
            col_lower = str(col).lower()
            if any(k in col_lower for k in ["법인", "회사", "company", "corp"]):
                company_col = col
            if any(k in col_lower for k in ["사업자", "등록번호", "biz"]):
                biz_num_col = col

        if not company_col:
            company_col = df.columns[0]

        return df, company_col, biz_num_col, header_row, sheet_name

    except Exception as e:
        print(f"[ERROR] 임원지분현황 파일 로드 실패: {e}")
        raise


def filter_data_by_manager(df, company_col, biz_num_col, manager_companies):
    """담당자별 데이터 필터링"""
    print(f"\n[3] 담당자별 데이터 필터링 수행")
    manager_data = {}

    for manager, companies in manager_companies.items():
        target_names = {c["company_name"] for c in companies}
        target_biz_nums = {c["biz_number"] for c in companies if c["biz_number"]}

        # 필터링 함수
        def check_row(row):
            # 1. 법인명 비교
            row_comp = normalize_company_name(row[company_col])
            # 정확히 일치하거나, 공백 제거 후 일치하는 경우
            if row_comp in target_names:
                return True
            for t_name in target_names:
                if row_comp == t_name:  # 이미 normalize 되어있으므로 단순비교
                    return True

            # 2. 사업자번호 비교 (있다면)
            if biz_num_col and pd.notna(row[biz_num_col]):
                row_biz = normalize_biz_number(row[biz_num_col])
                if row_biz in target_biz_nums:
                    return True
            return False

        mask = df.apply(check_row, axis=1)
        filtered_df = df[mask].copy()
        manager_data[manager] = filtered_df

    return manager_data


def copy_cell_style(source_cell, target_cell):
    """단일 셀 스타일 복사"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def copy_sheet_styles_full(
    source_wb, source_sheet_name, target_wb, target_sheet_name, header_row
):
    """
    원본 시트의 스타일을 타겟 시트에 복사 (헤더 + 데이터 영역 포함)
    """
    try:
        source_ws = source_wb[source_sheet_name]
        target_ws = target_wb[target_sheet_name]

        # 1. 열 너비 복사
        for col_idx in range(1, source_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in source_ws.column_dimensions:
                target_ws.column_dimensions[col_letter].width = (
                    source_ws.column_dimensions[col_letter].width
                )

        # 2. 헤더 행 스타일 복사
        # 헤더가 있는 행까지의 모든 스타일을 복사합니다.
        for row in range(1, header_row + 2):
            for col in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_ws.cell(row=row, column=col)
                copy_cell_style(source_cell, target_cell)

        # 3. 데이터 행 스타일 복사 (중요)
        # 원본 데이터의 첫 번째 행(header_row + 2)의 스타일을 샘플로 추출하여
        # 타겟의 모든 데이터 행에 적용합니다.

        source_data_start_row = header_row + 2
        target_max_row = target_ws.max_row

        # 원본에 데이터가 있는지 확인
        if source_ws.max_row >= source_data_start_row:
            # 컬럼별 스타일 샘플링
            col_styles = {}
            for col in range(1, source_ws.max_column + 1):
                col_styles[col] = source_ws.cell(row=source_data_start_row, column=col)

            # 타겟 데이터 전체에 적용
            for row in range(header_row + 2, target_max_row + 1):
                for col in range(1, source_ws.max_column + 1):
                    target_cell = target_ws.cell(row=row, column=col)
                    if col in col_styles:
                        copy_cell_style(col_styles[col], target_cell)

    except Exception as e:
        print(f"    [경고] 스타일 복사 중 오류 (일부 스타일 누락 가능): {e}")


def main():
    print("=" * 70)
    print("담당자별 임원지분현황 배포 파일 생성 프로그램")
    print("=" * 70)

    # 1. 경로 설정 (프로젝트 루트 디렉토리 기준)
    script_dir = Path(__file__).parent.parent
    output_dir = script_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    manager_file = script_dir / "smer" / "담당자정보_20251209.xlsx"
    stock_file = script_dir / "smer" / "임원지분현황.xlsx"

    # 파일 검증
    if not manager_file.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {manager_file}")
        return
    if not stock_file.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {stock_file}")
        return

    try:
        # 2. 데이터 로드 및 처리
        manager_companies, _ = load_manager_info(manager_file)
        stock_df, company_col, biz_num_col, header_row, src_sheet = (
            load_stock_holding_data(stock_file)
        )

        if not manager_companies:
            print("[ERROR] 처리할 담당자 정보가 없습니다.")
            return

        manager_data = filter_data_by_manager(
            stock_df, company_col, biz_num_col, manager_companies
        )

        # 3. 엑셀 파일 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"임원지분현황_담당자별배포_{timestamp}.xlsx"
        print(f"\n[4] 파일 생성 시작: {output_file.name}")

        # Pandas로 데이터 쓰기
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            has_data = False
            for manager, df in manager_data.items():
                if len(df) > 0:
                    sheet_name = manager[:30]  # 엑셀 시트명 길이 제한
                    # 헤더 포함하여 저장
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                    print(f"    - 시트 작성: {sheet_name} ({len(df)}건)")
                    has_data = True

            if not has_data:
                print("    [알림] 매칭된 데이터가 없습니다.")
                pd.DataFrame().to_excel(writer, sheet_name="데이터없음")

        # 4. 서식(스타일) 적용
        print("\n[5] 원본 서식 적용 중...")
        source_wb = load_workbook(stock_file)
        target_wb = load_workbook(output_file)

        for manager, df in manager_data.items():
            if len(df) > 0:
                sheet_name = manager[:30]
                if sheet_name in target_wb.sheetnames:
                    copy_sheet_styles_full(
                        source_wb, src_sheet, target_wb, sheet_name, header_row
                    )

        target_wb.save(output_file)
        target_wb.close()
        source_wb.close()

        print(f"\n[완료] 파일이 성공적으로 생성되었습니다: {output_file}")

    except Exception as e:
        print(f"\n[CRITICAL ERROR] 프로그램 실행 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
