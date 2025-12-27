# -*- coding: utf-8 -*-
"""
담당자별 임원지분현황 배포 파일 생성 스크립트
- 담당자정보 파일에서 담당자별 법인 매핑 정보 추출
- 임원지분현황 파일을 담당자별로 시트 분리
- 각 담당자가 담당하는 법인의 임원만 추출하여 시트 작성
- 각 시트는 원본 양식과 동일하게 구성
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys


def normalize_company_name(name):
  """회사명 정규화: 공백 제거, 앞뒤 공백 제거"""
  if pd.isna(name):
    return ""
  s = str(name).strip()
  # 여러 공백을 하나로
  s = ' '.join(s.split())
  return s


def normalize_biz_number(num):
  """사업자번호 정규화: 하이픈 제거, 소수점 제거"""
  if pd.isna(num):
    return ""
  s = str(num).replace("-", "").strip()
  if '.' in s:
    s = s.split('.')[0]
  return s


def load_manager_info(file_path):
  """
  담당자정보 파일 로드
  파일 구조를 확인하여 담당자와 법인 매핑 정보 추출
  """
  print(f"[1] 담당자정보 파일 로드 중: {file_path}")
  
  try:
    # 먼저 시트 이름 확인
    xl = pd.ExcelFile(file_path)
    print(f"    시트 목록: {xl.sheet_names}")
    
    # 첫 번째 시트 사용
    sheet_name = xl.sheet_names[0]
    print(f"    사용 시트: {sheet_name}")
    
    # 데이터 로드 (헤더 포함)
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    print(f"    컬럼 목록: {list(df.columns)}")
    print(f"    데이터 행 수: {len(df)}")
    
    # 담당자와 법인 정보가 있는 컬럼 찾기
    manager_col = None
    company_col = None
    biz_num_col = None
    
    for col in df.columns:
      col_lower = str(col).lower()
      if '담당자' in col_lower or 'manager' in col_lower or '담당' in col_lower:
        manager_col = col
      if '법인' in col_lower or '회사' in col_lower or 'company' in col_lower:
        company_col = col
      if '사업자' in col_lower or 'biz' in col_lower or '등록번호' in col_lower:
        biz_num_col = col
    
    if not manager_col or not company_col:
      print("\n[ERROR] 필수 컬럼을 찾을 수 없습니다.")
      print("사용 가능한 컬럼:")
      for i, col in enumerate(df.columns):
        print(f"  {i}: {col}")
      return None, None
    
    print(f"\n    사용 컬럼:")
    print(f"      담당자: {manager_col}")
    print(f"      법인명: {company_col}")
    if biz_num_col:
      print(f"      사업자번호: {biz_num_col}")
    
    # 담당자별 법인 그룹화
    manager_companies = {}
    for _, row in df.iterrows():
      manager = str(row[manager_col]).strip() if pd.notna(row[manager_col]) else ""
      company = normalize_company_name(row[company_col])
      
      if manager and company:
        if manager not in manager_companies:
          manager_companies[manager] = []
        
        company_info = {
          'company_name': company,
          'biz_number': normalize_biz_number(row[biz_num_col]) if biz_num_col and pd.notna(row[biz_num_col]) else ""
        }
        
        # 중복 제거
        if company_info not in manager_companies[manager]:
          manager_companies[manager].append(company_info)
    
    print(f"\n    담당자별 법인 수:")
    for manager, companies in manager_companies.items():
      print(f"      {manager}: {len(companies)}개 법인")
      for comp in companies:
        print(f"        - {comp['company_name']}")
    
    return manager_companies, df
  
  except Exception as e:
    print(f"[ERROR] 담당자정보 파일 로드 실패: {e}")
    raise


def load_stock_holding_data(file_path):
  """
  임원지분현황 파일 로드 - pandas로 전체 데이터 로드
  """
  print(f"\n[2] 임원지분현황 파일 로드 중: {file_path}")
  
  try:
    # 시트 이름 확인
    xl = pd.ExcelFile(file_path)
    print(f"    시트 목록: {xl.sheet_names}")
    
    # 첫 번째 시트 사용
    sheet_name = xl.sheet_names[0]
    print(f"    사용 시트: {sheet_name}")
    
    # 헤더 없이 전체 로드 (헤더 행 수를 확인하기 위해)
    df_no_header = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # 헤더 행 찾기 (일반적으로 처음 몇 행)
    header_row = 0
    for idx in range(min(10, len(df_no_header))):
      row_values = [str(v).lower() for v in df_no_header.iloc[idx] if pd.notna(v)]
      row_str = " ".join(row_values)
      if any(keyword in row_str for keyword in ['법인', '회사', 'company', '기업']):
        header_row = idx
        break
    
    print(f"    헤더 행: {header_row}")
    
    # 헤더를 포함하여 다시 로드
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    
    print(f"    컬럼 목록: {list(df.columns)}")
    print(f"    데이터 행 수: {len(df)}")
    
    # 법인명 컬럼 찾기
    company_col = None
    biz_num_col = None
    
    for col in df.columns:
      col_lower = str(col).lower()
      if any(keyword in col_lower for keyword in ['법인', '회사', 'company', '기업', 'corp']):
        company_col = col
      if any(keyword in col_lower for keyword in ['사업자', '등록번호', 'biz', 'business', '번호']):
        biz_num_col = col
    
    print(f"    법인명 컬럼: {company_col}")
    print(f"    사업자번호 컬럼: {biz_num_col}")
    
    if not company_col:
      print("\n[경고] 법인명 컬럼을 찾을 수 없습니다. 첫 번째 컬럼을 사용합니다.")
      company_col = df.columns[0]
    
    return df, company_col, biz_num_col, header_row, sheet_name
  
  except Exception as e:
    print(f"[ERROR] 임원지분현황 파일 로드 실패: {e}")
    raise


def filter_data_by_manager(df, company_col, biz_num_col, manager_companies):
  """
  담당자별로 데이터 필터링
  각 담당자가 담당하는 법인의 임원만 정확히 필터링
  """
  print(f"\n[3] 담당자별 데이터 필터링 중...")
  
  manager_data = {}
  
  for manager, companies in manager_companies.items():
    # 담당 법인명과 사업자번호 정규화
    company_names = {normalize_company_name(c['company_name']) for c in companies}
    biz_numbers = {c['biz_number'] for c in companies if c['biz_number']}
    
    print(f"\n    [{manager}] 담당 법인 ({len(companies)}개):")
    for comp in companies:
      print(f"      - {comp['company_name']}")
    
    # 필터 조건 생성
    conditions = []
    
    # 법인명으로 정확히 필터링
    if company_col:
      def match_company(x):
        if pd.isna(x):
          return False
        normalized_x = normalize_company_name(x)
        
        # 정확 일치 확인
        if normalized_x in company_names:
          return True
        
        # 공백 차이만 있는 경우 처리
        for name in company_names:
          # 공백 제거 후 비교
          if normalized_x.replace(' ', '') == name.replace(' ', ''):
            return True
          if normalized_x.replace(' ', '') == name or name.replace(' ', '') == normalized_x:
            return True
        
        return False
      
      company_condition = df[company_col].apply(match_company)
      conditions.append(company_condition)
    
    # 사업자번호로 필터링 (정확 일치만)
    if biz_num_col:
      biz_condition = df[biz_num_col].apply(
        lambda x: normalize_biz_number(x) in biz_numbers if pd.notna(x) else False
      )
      conditions.append(biz_condition)
    
    # 조건 결합 (OR 조건 - 법인명 또는 사업자번호 중 하나라도 일치하면 포함)
    if conditions:
      combined_condition = conditions[0]
      for cond in conditions[1:]:
        combined_condition = combined_condition | cond
      
      filtered_df = df[combined_condition].copy()
    else:
      filtered_df = pd.DataFrame()
    
    manager_data[manager] = filtered_df
    
    # 필터링 결과 확인
    if len(filtered_df) > 0 and company_col:
      matched_companies = filtered_df[company_col].apply(normalize_company_name).unique()
      print(f"    매칭된 법인: {list(matched_companies)}")
      print(f"    매칭된 행 수: {len(filtered_df)}개")
    else:
      print(f"    매칭된 행 수: 0개 (데이터 없음)")
  
  return manager_data


def copy_sheet_styles(source_wb, source_sheet_name, target_wb, target_sheet_name, header_row):
  """
  원본 시트의 스타일을 타겟 시트에 복사
  """
  try:
    source_ws = source_wb[source_sheet_name]
    target_ws = target_wb[target_sheet_name]
    
    # 열 너비 복사
    for col_idx in range(1, min(source_ws.max_column + 1, target_ws.max_column + 1)):
      col_letter = get_column_letter(col_idx)
      if col_letter in source_ws.column_dimensions:
        source_width = source_ws.column_dimensions[col_letter].width
        if source_width:
          target_ws.column_dimensions[col_letter].width = source_width
    
    # 헤더 행 스타일 복사
    for row_idx in range(1, header_row + 2):  # 헤더 행 + 1
      for col_idx in range(1, min(source_ws.max_column + 1, target_ws.max_column + 1)):
        try:
          source_cell = source_ws.cell(row=row_idx, column=col_idx)
          target_cell = target_ws.cell(row=row_idx, column=col_idx)
          
          if source_cell.has_style:
            target_cell.font = source_cell.font
            target_cell.alignment = source_cell.alignment
            target_cell.border = source_cell.border
            target_cell.fill = source_cell.fill
            target_cell.number_format = source_cell.number_format
        except:
          pass
  except Exception as e:
    print(f"    [경고] 스타일 복사 중 오류 (무시): {e}")


def main():
  """메인 함수"""
  print("=" * 70)
  print("담당자별 임원지분현황 배포 파일 생성 프로그램")
  print("=" * 70)
  
  # 파일 경로 설정
  script_dir = Path(__file__).parent
  
  manager_file = script_dir / '담당자정보_20251209.xlsx'
  stock_file = script_dir / '임원지분현황.xlsx'
  output_file = script_dir / f'임원지분현황_담당자별배포_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
  
  # 파일 존재 확인
  if not manager_file.exists():
    print(f"\n[ERROR] 담당자정보 파일을 찾을 수 없습니다: {manager_file}")
    return
  
  if not stock_file.exists():
    print(f"\n[ERROR] 임원지분현황 파일을 찾을 수 없습니다: {stock_file}")
    return
  
  try:
    # 1. 담당자정보 로드
    manager_companies, manager_df = load_manager_info(manager_file)
    
    if not manager_companies:
      print("\n[ERROR] 담당자별 법인 정보를 찾을 수 없습니다.")
      return
    
    # 2. 임원지분현황 파일 로드
    stock_df, company_col, biz_num_col, header_row, source_sheet_name = load_stock_holding_data(stock_file)
    
    # 3. 담당자별 데이터 필터링
    manager_data = filter_data_by_manager(stock_df, company_col, biz_num_col, manager_companies)
    
    # 4. 원본 파일의 스타일 정보 로드
    print(f"\n[4] 원본 파일 스타일 로드 중...")
    source_wb = load_workbook(stock_file, data_only=False)
    
    # 5. 배포용 파일 생성 - pandas ExcelWriter 사용
    print(f"\n[5] 배포용 파일 생성 중: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
      created_sheets = 0
      
      for manager, filtered_df in manager_data.items():
        if len(filtered_df) > 0:
          # 시트 이름은 31자 제한
          sheet_name = manager[:31] if len(manager) > 31 else manager
          
          # 헤더 없이 저장 (이미 헤더가 포함된 데이터프레임)
          filtered_df.to_excel(
            writer, 
            sheet_name=sheet_name, 
            index=False,
            header=True
          )
          
          print(f"    시트 생성: {sheet_name} ({len(filtered_df)}개 행)")
          created_sheets += 1
      
      # 최소 하나의 시트가 있어야 함
      if created_sheets == 0:
        print("\n[경고] 데이터가 없어 빈 시트를 생성합니다.")
        first_manager = list(manager_data.keys())[0] if manager_data else "데이터없음"
        sheet_name = first_manager[:31] if len(first_manager) > 31 else first_manager
        
        # 헤더만 있는 빈 데이터프레임 생성
        empty_df = pd.DataFrame(columns=stock_df.columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
        created_sheets = 1
    
    # 스타일 복사
    print(f"\n[6] 스타일 복사 중...")
    target_wb = load_workbook(output_file)
    
    for manager, filtered_df in manager_data.items():
      if len(filtered_df) > 0:
        sheet_name = manager[:31] if len(manager) > 31 else manager
        if sheet_name in target_wb.sheetnames:
          copy_sheet_styles(source_wb, source_sheet_name, target_wb, sheet_name, header_row)
    
    # 파일 저장
    target_wb.save(output_file)
    target_wb.close()
    source_wb.close()
    
    print(f"\n[OK] 배포용 파일 저장 완료: {output_file}")
    
    # 요약 정보 출력
    print("\n" + "=" * 70)
    print("생성된 시트 요약")
    print("=" * 70)
    for manager, filtered_df in manager_data.items():
      print(f"  {manager}: {len(filtered_df)}개 행, {len(manager_companies[manager])}개 법인")
    
    print("\n프로그램 완료!")
  
  except Exception as e:
    print(f"\n[ERROR] 오류 발생: {e}")
    import traceback
    traceback.print_exc()
    return


if __name__ == '__main__':
  main()
